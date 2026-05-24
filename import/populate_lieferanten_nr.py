"""
Populate skus.lieferanten_nr from "B2B Artikel mit VK.xlsx" (Zeile1 only).

Usage:
  SUPABASE_URL=... SUPABASE_SERVICE_KEY=... uv run --with openpyxl --with supabase python3 populate_lieferanten_nr.py
"""

import os
import openpyxl
from supabase import create_client

XLSX = os.path.join(os.path.dirname(__file__), "B2B Artikel mit VK.xlsx")

url = os.environ["SUPABASE_URL"]
key = os.environ["SUPABASE_SERVICE_KEY"]
sb = create_client(url, key)

wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
ws = wb.active

# Detect header row
header = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]

# Artikel-Nr is stored as "Teilenummer" in this file
col_artikel = header.index("Teilenummer")

# Find Zeile1 column
col_zeile1 = None
for i, h in enumerate(header):
    if "Zeile1" in h:
        col_zeile1 = i
        break
if col_zeile1 is None:
    raise RuntimeError(f"Could not find Zeile1 column. Headers: {header}")

print(f"Artikel-Nr col: {col_artikel} ({header[col_artikel]})")
print(f"Zeile1 col:     {col_zeile1} ({header[col_zeile1]})")

rows_to_update: list[tuple[str, str]] = []
for row in ws.iter_rows(min_row=2, values_only=True):
    artikel_nr = row[col_artikel]
    zeile1 = row[col_zeile1]
    if not artikel_nr or not zeile1:
        continue
    rows_to_update.append((str(artikel_nr).strip(), str(zeile1).strip()))

print(f"Rows with data: {len(rows_to_update)}")

updated = 0
not_found: list[str] = []

for artikel_nr, lieferanten_nr in rows_to_update:
    res = sb.table("skus").update({"lieferanten_nr": lieferanten_nr}).eq("artikel_nr", artikel_nr).execute()
    if res.data:
        updated += len(res.data)
    else:
        not_found.append(artikel_nr)

print(f"Updated: {updated} SKU rows")
if not_found:
    print(f"Not found in DB ({len(not_found)}): {not_found[:20]}")
else:
    print("All artikel_nrs matched.")

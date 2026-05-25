"""
Import material suitability data from B2B_Artikeltexte_komplett_2026_gefuellt.xlsx
into product_materials for products that currently have no material data.

Join chain: Artikeltexte.Teilenummer → skus.artikel_nr → skus.product_id → product

Safety: products that already have product_materials rows are skipped entirely.

Usage:
  SUPABASE_URL=... SUPABASE_SERVICE_ROLE_KEY=... \\
    uv run --with supabase --with openpyxl \\
    python3 import/import_materials_artikeltexte.py [--dry-run]
"""

import os, sys, re
from collections import defaultdict
import openpyxl
from supabase import create_client

XLSX = os.path.join(os.path.dirname(__file__), "B2B_Artikeltexte_komplett_2026_gefuellt (2).xlsx")
DRY_RUN = "--dry-run" in sys.argv

VALID_SUITABILITY = {"sehr gut geeignet", "gut geeignet", "geeignet", "nicht geeignet"}

url = os.environ["SUPABASE_URL"]
key = os.environ["SUPABASE_SERVICE_ROLE_KEY"]
sb = create_client(url, key)


def parse_material_text(text: str) -> list[tuple[str, str]]:
    """Parse 'Mat A (suitability), Mat B (suitability)' → [(name, suitability), ...]"""
    results = []
    if not text:
        return results
    # Split on "), " but be careful of parens inside material names
    parts = re.split(r"\),\s*", text.strip().rstrip(")"))
    for part in parts:
        part = part.strip()
        if not part:
            continue
        # Find last " (" to split name from suitability
        m = re.match(r"^(.+)\s+\(([^)]+)\)?\s*$", part)
        if not m:
            # Try simpler: last " (" split
            idx = part.rfind(" (")
            if idx == -1:
                continue
            name = part[:idx].strip()
            suitability = part[idx + 2:].strip().rstrip(")")
        else:
            name = m.group(1).strip()
            suitability = m.group(2).strip()
        if suitability in VALID_SUITABILITY:
            results.append((name, suitability))
    return results


# --- Parse Excel (both sheets) ---
# teil_materials: Teilenummer → list of (material_name, suitability)
teil_materials: dict[str, list[tuple[str, str]]] = {}

wb = openpyxl.load_workbook(XLSX, read_only=True, data_only=True)
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    rows_iter = ws.iter_rows(min_row=1, values_only=True)
    header = next(rows_iter)

    for row in rows_iter:
        teilnr = str(row[0]).strip() if row[0] else ""
        if not teilnr or teilnr == "None":
            continue
        if teilnr in teil_materials:
            continue  # already got data for this Teilenummer from first sheet

        # Scan Überschrift/Text pairs: columns 1-2, 3-4, 5-6
        mat_text = None
        for ub_idx, txt_idx in [(1, 2), (3, 4), (5, 6)]:
            if ub_idx >= len(row) or txt_idx >= len(row):
                break
            ub = str(row[ub_idx]).strip() if row[ub_idx] else ""
            if ub == "Materialien":
                mat_text = str(row[txt_idx]).strip() if row[txt_idx] else ""
                break

        if mat_text and mat_text != "None":
            parsed = parse_material_text(mat_text)
            if parsed:
                teil_materials[teilnr] = parsed

print(f"Artikeltexte rows with material data: {len(teil_materials)}")

# --- Fetch DB ---
all_skus = sb.table("skus").select("artikel_nr,product_id").execute().data
# artikel_nr → product_id (may be null if SKU not linked)
sku_to_product: dict[str, str] = {
    s["artikel_nr"]: s["product_id"]
    for s in all_skus
    if s.get("product_id") and s.get("artikel_nr")
}

all_products = sb.table("products").select("id,name,slug").execute().data
prod_info: dict[str, dict] = {p["id"]: p for p in all_products}

existing_materials = sb.table("product_materials").select("product_id,material_name,sort_order").execute().data
# product_id → set of material names already in DB
existing_by_product: dict[str, set[str]] = defaultdict(set)
max_sort_by_product: dict[str, int] = defaultdict(int)
for m in existing_materials:
    pid = m["product_id"]
    existing_by_product[pid].add(m["material_name"])
    if (m["sort_order"] or 0) > max_sort_by_product[pid]:
        max_sort_by_product[pid] = m["sort_order"] or 0

existing_material_types = {
    r["name"] for r in sb.table("material_types").select("name").execute().data
}

# --- Group materials by product_id ---
# For products with multiple SKUs, union their materials
product_materials_map: dict[str, dict[str, str]] = defaultdict(dict)  # pid → {mat_name: suitability}
skus_not_found: list[str] = []

for teilnr, materials in teil_materials.items():
    product_id = sku_to_product.get(teilnr)
    if not product_id:
        skus_not_found.append(teilnr)
        continue
    for name, suitability in materials:
        if name not in product_materials_map[product_id]:
            product_materials_map[product_id][name] = suitability

# --- Insert ---
products_updated = 0
products_skipped_existing = 0
total_materials_inserted = 0
new_material_types_added = []

for product_id, mat_dict in sorted(product_materials_map.items(), key=lambda x: prod_info.get(x[0], {}).get("slug", "")):
    prod = prod_info.get(product_id, {})
    slug = prod.get("slug", product_id)

    if existing_by_product[product_id]:
        products_skipped_existing += 1
        continue

    rows_to_insert = []
    next_sort = max_sort_by_product[product_id]

    for mat_name, suitability in mat_dict.items():
        # Ensure material_type exists (FK constraint)
        if mat_name not in existing_material_types:
            if not DRY_RUN:
                sb.table("material_types").insert({"name": mat_name}).execute()
            existing_material_types.add(mat_name)
            new_material_types_added.append(mat_name)

        rows_to_insert.append({
            "product_id": product_id,
            "material_name": mat_name,
            "suitability": suitability,
            "sort_order": next_sort,
        })
        next_sort += 1

    if rows_to_insert:
        if not DRY_RUN:
            sb.table("product_materials").insert(rows_to_insert).execute()
        total_materials_inserted += len(rows_to_insert)
        verb = "[DRY-RUN]" if DRY_RUN else "[INSERTED]"
        print(f"{verb} {slug}: {len(rows_to_insert)} materials")
        products_updated += 1

# --- Report ---
print(f"\n--- Summary ---")
print(f"Products updated:                {products_updated}")
print(f"Material entries inserted:       {total_materials_inserted}")
print(f"Products skipped (had data):     {products_skipped_existing}")
print(f"Teilenummern not in DB skus:     {len(skus_not_found)}")
if new_material_types_added:
    print(f"New material_types added ({len(new_material_types_added)}): {new_material_types_added}")
if DRY_RUN:
    print("\n[DRY-RUN] No data was written.")
